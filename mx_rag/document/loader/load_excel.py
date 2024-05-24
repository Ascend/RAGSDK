# -*- coding: utf-8 -*-
# Copyright (c) Huawei Technologies Co., Ltd. 2024. All rights reserved.

from openpyxl import load_workbook
import xlrd
from loguru import logger
from mx_rag.utils import file_check
from mx_rag.utils.file_check import FileBrokenException

OPENPYXL_EXTENSION = (".xlsx",)
XLRD_EXTENSION = (".xls",)


class ExcelLoader:
    def __init__(self, file_path, max_size_mb=100):
        self.file_path = file_path
        self.max_size_mb = max_size_mb

    @staticmethod
    def _cleanup_xlsx(worksheet):
        """
        ：返回：去掉左边与上边空白行列后的表单，左上角对其, 需要感知空cell
        """
        # 查找表格最左上角的cell位置
        start_row = 0
        start_col = 0
        for row in worksheet.iter_rows(values_only=True):
            if all(cell is None for cell in row):
                start_row += 1
        for col in worksheet.iter_cols(values_only=True):
            if all(cell is None for cell in col):
                start_col += 1
        # 删除空行和空列
        if start_row:
            worksheet.delete_rows(1, amount=start_row)
        if start_col:
            worksheet.delete_cols(1, amount=start_col)
        return worksheet

    @staticmethod
    def _table_location_xls(worksheet):
        """
        查找表格最左上角的cell位置
        """
        start_row = 0
        start_col = 0
        for i in range(worksheet.nrows):
            row = worksheet.row_values(i)
            if not sum(cell != "" for cell in row):
                start_row += 1
        for i in range(worksheet.ncols):
            col = worksheet.col_values(i)
            if not sum(cell != "" for cell in col):
                start_col += 1
        return start_row, start_col

    def load(self):
        """
        ：返回：逐行读取表,返回 string list
        """
        try:
            file_check.excel_file_check(self.file_path, self.max_size_mb * 1024 * 1024)
        except Exception as e:
            logger.error(e)
            return []
        # 判断文件种类：支持 xlsx 与 xls 格式
        if self.file_path.endswith(XLRD_EXTENSION):
            return self._load_xls()
        elif self.file_path.endswith(OPENPYXL_EXTENSION):
            return self._load_xlsx()
        else:
            raise TypeError(f"{self.file_path} file type is not correct")

    def _load_xls(self):
        try:
            wb = xlrd.open_workbook(self.file_path)
        except Exception as e:
            raise FileBrokenException(f"{self.file_path} is broken cannot load") from e
        res = []
        for i in range(wb.nsheets):
            ws = wb.sheet_by_index(i)
            start_row, start_col = self._table_location_xls(ws)
            if ws.nrows - start_row < 2:
                logger.info(f"In file {self.file_path} sheet *{ws.name}* is empty")
                continue
            title = ws.row_values(start_row)[start_col:]
            for line_ind in range(start_row + 1, ws.nrows):
                text_line = ""
                line_list = ws.row_values(line_ind)
                for ind, ti in enumerate(title):
                    ind += start_col
                    text_line += str(ti) + ": " + str(line_list[ind]) + "; "
                text_line += "--" + str(ws.name)
                res.append(text_line)
        logger.info(f"file {self.file_path} Loading completed")
        return res

    def _load_xlsx(self):
        try:
            wb = load_workbook(self.file_path)
        except Exception as e:
            raise FileBrokenException(f"{self.file_path} is broken cannot load") from e
        res = []
        for sheet_name in wb.sheetnames:
            ws_init = wb[sheet_name]
            ws = self._cleanup_xlsx(ws_init)
            rows = list(ws.rows)
            # 判断表单是否有标题+内容，默认至少两行
            if len(rows) < 2:
                logger.info(f"In file {self.file_path} sheet *{sheet_name}* is empty")
                continue
            title = list(rows[0])
            for line in list(rows[1:]):
                text_line = ""
                for ind, ti in enumerate(title):
                    text_line += str(ti.value) + ": " + str(line[ind].value) + "; "
                text_line += "--" + str(sheet_name)
                res.append(text_line)
        logger.info(f"file {self.file_path} Loading completed")
        return res
