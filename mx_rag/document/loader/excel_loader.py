# -*- coding: utf-8 -*-
# Copyright (c) Huawei Technologies Co., Ltd. 2024. All rights reserved.
import csv
from datetime import datetime, timedelta
from typing import (List, Iterator)
from openpyxl import load_workbook, Workbook
from openpyxl.cell import MergedCell
import xlrd
from loguru import logger
from langchain_core.documents import Document
from langchain_community.document_loaders.base import BaseLoader

from mx_rag.document.loader.base_loader import BaseLoader as mxBaseLoader
from mx_rag.utils import file_check


OPENPYXL_EXTENSION = (".xlsx",)
XLRD_EXTENSION = (".xls",)
CSV_EXTENSION = (".csv",)


class ExcelLoader(BaseLoader, mxBaseLoader):

    def __init__(self, file_path, line_sep="**;"):
        super().__init__(file_path)
        self.line_sep = str(line_sep)


    @staticmethod
    def _exceltime_to_datetime(exceltime):
        """
        将excel储存的时间格式转换为可读的格式
        :param exceltime: (0,1)区间的浮点数，表示时间在一天中的位置
        :return: 以 时：分 的格式返回字符串
        """
        base_date = datetime(1899, 12, 30)
        time = base_date + timedelta(exceltime)
        return time.strftime("%H:%M")

    @staticmethod
    def _get_xlsx_blank_rows_and_cols(worksheet):
        """
        功能：获取所有空行、空列的索引
        """
        row_count = 0
        col_count = 0

        null_rows = {}
        for row in worksheet.iter_rows(values_only=True):
            row_count += 1
            if all(not cell for cell in row):
                null_rows[row_count] = True

        blank_cols = {}
        for col in worksheet.iter_cols(values_only=True):
            col_count += 1
            if all(not cell for cell in col):
                blank_cols[col_count] = True

        return null_rows, blank_cols

    @staticmethod
    def _get_xls_blank_rows_and_cols(worksheet):
        """
        功能：获取所有空行、空列的索引
        """
        null_rows = {}
        blank_cols = {}

        for i in range(worksheet.nrows):
            row = worksheet.row_values(i)
            if all(not cell for cell in row):
                null_rows[i] = True

        for i in range(worksheet.ncols):
            col = worksheet.col_values(i)
            if all(not cell for cell in col):
                blank_cols[i] = True

        return null_rows, blank_cols

    @staticmethod
    def _parse_xlsx_cell(sheet: Workbook, row: int, col: int):
        """
        功能：读取xlsx cell值，如果是合并项，使用左上cell值替代
        """
        cell = sheet.cell(row=row, column=col)
        if not isinstance(cell, MergedCell):
            return cell.value

        for merged_range in sheet.merged_cells.ranges:
            if cell.coordinate in merged_range:
                # return the left top cell
                cell = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                return cell.value

        return cell.value

    @staticmethod
    def _parse_xls_cell(sheet, row: int, col: int):
        """
        功能：读取xls cell值，如果是合并项，使用左上cell值替代
        """
        cell = sheet.cell(rowx=row, colx=col)
        if cell.value:
            return cell.value

        for crange in sheet.merged_cells:
            rlo, rhi, clo, chi = crange
            if rlo <= row < rhi and clo <= col < chi:
                return sheet.cell(rowx=rlo, colx=clo).value

        return cell.value

    @staticmethod
    def _get_xlsx_first_not_blank_row_and_col(ws):
        """
        功能：读取xlsx 非空白首行、首列索引
        """
        first_row = 1
        first_col = 1
        for row in ws.iter_rows(values_only=True):
            if any(cell for cell in row):
                break
            first_row += 1

        for col in ws.iter_cols(values_only=True):
            if any(cell for cell in col):
                break
            first_col += 1

        return first_row, first_col

    @staticmethod
    def _get_xls_first_not_blank_row_and_col(ws):
        """
        功能：获取xls 非空白首行、首列索引
        """
        first_row = 0
        first_col = 0
        for i in range(ws.nrows):
            if any(cell for cell in ws.row_values(i)):
                first_row = i
                break

        for i in range(ws.ncols):
            if any(cell for cell in ws.col_values(i)):
                first_col = i
                break

        return first_row, first_col

    @staticmethod
    def _load_csv_line(row, headers):
        text_line = ""
        for ind, ti in enumerate(headers):
            if not str(ti):
                ti = "None"
            if not str(row[ind]):
                row[ind] = "None"
            text_line += str(ti) + ":" + str(row[ind]) + ";"
        return text_line

    def lazy_load(self) -> Iterator[Document]:
        """
        ：返回：逐行读取表,返回 string list
        """
        try:
            file_check.excel_file_check(self.file_path, self.MAX_SIZE)
        except Exception as e:
            logger.error(e)
            return iter([])
        # 判断文件种类：支持 xlsx 与 xls 格式
        if self.file_path.endswith(XLRD_EXTENSION):
            return self._load_xls()
        elif self.file_path.endswith(OPENPYXL_EXTENSION):
            if self._is_zip_bomb():
                return iter([])
            else:
                return self._load_xlsx()
        elif self.file_path.endswith(CSV_EXTENSION):
            return self._load_csv()
        else:
            logger.error(f"{self.file_path} file type is not one of (csv, xlsx, xls).")
            return iter([])

    def _get_xlsx_title(self, ws, first_row, first_col):
        title = []
        for col in range(first_col, ws.max_column + 1):
            ti = str(self._parse_xlsx_cell(ws, first_row, col))
            title.append(ti if ti else "None")

        return title

    def _get_xls_title(self, ws, first_row, first_col):
        title = []
        for col in range(first_col, ws.ncols):
            ti = str(self._parse_xls_cell(ws, first_row, col))
            title.append(ti if ti else "None")

        return title

    def _load_xlsx_one_sheet(self, ws):
        """
        功能：读取一个xlsx表单的值，每行值以title:value;title:value....的格式，行与行之间通过self.line_sep分隔
        """
        content = ""
        blank_rows, blank_cols = self._get_xlsx_blank_rows_and_cols(ws)

        # 获取有效第一行,列的索引
        first_row, first_col = self._get_xlsx_first_not_blank_row_and_col(ws)

        # 判断表单是否有标题+内容，默认至少两行有效行
        if ws.max_row - len(blank_rows.keys()) < 2:
            return content

        # 获取标题列表
        title = self._get_xlsx_title(ws, first_row, first_col)

        column_end = ws.max_column + 1
        for row_index in range(first_row + 1, ws.max_row + 1):
            # 空行无数据，不解析
            if row_index in blank_rows.keys():
                continue

            text_line = ""
            for col_index in range(1, column_end):
                # 空列无数据，不解析
                if col_index in blank_cols.keys():
                    continue

                val = self._parse_xlsx_cell(ws, row_index, col_index)
                ti = title[col_index - first_col]
                text_line += str(ti) + ":" + str(val) + ";"

            content += text_line + self.line_sep

        return content

    def _load_xls_one_sheet(self, ws):
        """
        功能：读取一个xls表单的值，每行值以title:value;title:value....的格式，行与行之间通过self.line_sep分隔
        """
        content = ""
        blank_rows, blank_cols = self._get_xls_blank_rows_and_cols(ws)
        # 获取有效第一行,列的索引
        first_row, first_col = self._get_xls_first_not_blank_row_and_col(ws)

        # 判断表单是否有标题+内容，默认至少两行有效行
        if ws.nrows - len(blank_rows.keys()) < 2:
            return content

        # 获取标题列表
        title = self._get_xls_title(ws, first_row, first_col)
        ncols = ws.ncols
        for row_index in range(first_row + 1, ws.nrows):
            # 空行无数据，不解析
            if row_index in blank_rows.keys():
                continue

            text_line = ""
            for col_index in range(ncols):
                # 空列无数据，不解析
                if col_index in blank_cols.keys():
                    continue
                val = self._parse_xls_cell(ws, row_index, col_index)
                ti = title[col_index - first_col]
                if ti in ["time"] and 0 <= float(val) <= 1:
                    text_line += str(ti) + ":" + str(self._exceltime_to_datetime(float(val))) + ";"
                else:
                    text_line += str(ti) + ":" + str(val) + ";"

            content += text_line + self.line_sep

        return content

    def _load_xls(self):
        wb = xlrd.open_workbook(self.file_path, formatting_info=True)
        if wb.nsheets > self.MAX_PAGE_NUM:
            logger.error(f"file {self.file_path} sheets number more than limit")
            yield Document(page_content='')
        for i in range(wb.nsheets):
            ws = wb.sheet_by_index(i)
            content = self._load_xls_one_sheet(ws)
            if not content:
                logger.info(f"In file [{self.file_path}] sheet [{ws.name}] is empty")
                continue
            yield Document(page_content=content, metadata={"source": self.file_path, "sheet": ws.name})

        logger.info(f"file {self.file_path} Loading completed")

    def _load_xlsx(self):
        wb = load_workbook(self.file_path, data_only=True)
        if len(wb.sheetnames) > self.MAX_PAGE_NUM:
            logger.error(f"file {self.file_path} sheets number more than limit")
            yield Document(page_content='')
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            content = self._load_xlsx_one_sheet(ws)
            if not content:
                logger.info(f"In file [{self.file_path}] sheet [{sheet_name}] is empty")
                continue
            yield Document(page_content=content, metadata={"source": self.file_path, "sheet": sheet_name})
        logger.info(f"file {self.file_path} Loading completed")

    def _load_csv_lines(self, reader, headers):
        content = ""
        for row in reader:
            if len(row) > 0:
                text_line = self._load_csv_line(row, headers)
                content += text_line + self.line_sep
            else:
                break
        return content

    def _load_csv(self):
        content = ""
        try:
            with open(self.file_path, mode="r", encoding="utf-8-sig") as file:
                reader = csv.reader(file)
                headers = next(reader)  # 读取第一行标题
                content = self._load_csv_lines(reader, headers)
        except Exception as e:
            logger.error(e)
            yield Document(page_content='')
        if content:
            yield Document(page_content=content, metadata={"source": self.file_path})
        else:
            logger.info(f"file {self.file_path} is empty")
            yield Document(page_content='')

